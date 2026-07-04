"""
wq_bp.py — HydroCore Water Quality Blueprint
=============================================
Sampling stations, indicator readings, compliance checks, trend analysis,
CCME-WQI, ion balance, Piper/Stiff diagrams, Mann-Kendall, LSI, eutrophication,
AMD assessment, seasonal analysis, pollution load, Blue Drop scoring, and alerts.
All routes are org-scoped via JWT claims and require authentication.
"""

import io
import json
import logging
from datetime import datetime, date

import pandas as pd
from flask import Blueprint, g, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from auth import verify_jwt_claims
from wq_calculations import (
    compute_ccme_wqi, compute_ion_balance_error, compute_mann_kendall,
    compute_lsi, compute_trophic_state_index, compute_piper_data,
    compute_stiff_data, compute_blue_drop_score,
    compute_sar, compute_irrigation_suitability, compute_livestock_suitability,
    compute_aquaculture_suitability, classify_hardness, compute_amd_flags,
    compute_temporal_stability, diagnose_contamination_source,
    explain_diagnostic_finding, compute_ecosystem_health_score,
)
from wq_alerts import dispatch_alert

logger = logging.getLogger(__name__)

wq_bp = Blueprint('wq', __name__)

_db = None


def init_wq(db):
    global _db
    _db = db


def _q(sql, params=None, fetch=True, commit=False):
    return _db.execute_query(sql, params, fetch=fetch, commit=commit)


# ── Auth guard ────────────────────────────────────────────────────────────────

def require_wq_auth(roles=None):
    """Decorator: require valid JWT; optionally restrict to given roles."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            claims, err = verify_jwt_claims(roles)
            if err:
                return err
            g.wq_user_id = int(get_jwt_identity())
            g.wq_org_id  = int(claims.get('org_id', 1))
            g.wq_role    = claims.get('role', 'viewer')
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_float(val):
    try:
        return float(val) if val is not None and str(val).strip() not in ('', 'nan', 'NaN', 'NaT') else None
    except (ValueError, TypeError):
        return None


_ASCII_SAFE_REPLACEMENTS = {
    '–': '-',   # en dash
    '—': '-',   # em dash
    'µ': 'u',   # micro sign
    '°': 'deg', # degree sign
    '‰': 'permil',
    '≤': '<=',
    '≥': '>=',
    '≈': '~',
}


def _ascii_safe(text):
    """
    Replace characters this DB connection's pyodbc SQL_WCHAR-as-utf8 decoding
    cannot safely round-trip (confirmed to corrupt or raise UnicodeDecodeError
    on read-back for NVARCHAR data containing them). Scoped to text this app
    persists into new diagnostic-engine columns — does not touch
    wq_calculations.py's AMD_FLAG_THRESHOLDS constant, which must stay
    byte-identical for analytics_amd()'s live (non-persisted) JSON response.
    """
    if text is None:
        return text
    for uc, ascii_c in _ASCII_SAFE_REPLACEMENTS.items():
        text = text.replace(uc, ascii_c)
    return text


def _compute_compliance(measured_val, lower_std, upper_std):
    """Return (is_compliant, flag) based on indicator thresholds (SANS 241)."""
    if measured_val is None:
        return None, None
    has_limit = (lower_std is not None) or (upper_std is not None)
    if not has_limit:
        return None, None
    if lower_std is not None and measured_val < lower_std:
        return False, 'BELOW'
    if upper_std is not None and measured_val > upper_std:
        return False, 'ABOVE'
    return True, None


def _hardness_class(hardness_val):
    """Map a Total Hardness reading (mg CaCO3/L) to a DWAF Vol 7 hardness band.
    Defaults to 'medium' when hardness wasn't measured for this reading, since
    that's the least extreme assumption rather than silently skipping the check."""
    if hardness_val is None:
        return 'medium'
    if hardness_val < 60:
        return 'soft'
    if hardness_val < 120:
        return 'medium'
    if hardness_val <= 180:
        return 'hard'
    return 'very_hard'


def _compute_twqr_compliance(measured_val, indicator, hardness_val, hardness_bands):
    """
    Return (is_compliant, flag) against DWAF Vol 7 (Aquatic Ecosystems) TWQR —
    same shape as _compute_compliance so both can be stored/read symmetrically.

    hardness_bands: {indicator_id: {hardness_class: (twqr_upper, cev, aev)}},
    only populated for the 3 hardness-dependent indicators (Cd, Cu, Pb).
    """
    if measured_val is None:
        return None, None
    basis = indicator.get('twqr_basis')
    if basis in (None, 'not_applicable'):
        return None, None

    if basis == 'hardness_dependent':
        band = _hardness_class(hardness_val)
        row = hardness_bands.get(indicator['id'], {}).get(band)
        if not row:
            return None, None
        twqr_upper, _cev, _aev = row
        return (False, 'ABOVE') if measured_val > twqr_upper else (True, None)

    # 'fixed' and 'trophic_band' both reduce to the same lower/upper check —
    # trophic_band's twqr_upper is just the eutrophic-boundary cutover.
    lower, upper = indicator.get('twqr_lower'), indicator.get('twqr_upper')
    if lower is None and upper is None:
        return None, None
    if lower is not None and measured_val < lower:
        return False, 'BELOW'
    if upper is not None and measured_val > upper:
        return False, 'ABOVE'
    return True, None


def _fetch_indicators(org_id):
    """Return list of indicator dicts for a tenant (WQ_Indicators is tenant-owned)."""
    rows = _q(
        "SELECT indicator_id, indicator_code, indicator_name, unit, lower_std, upper_std, std_reference, "
        "display_order, risk_class, twqr_lower, twqr_upper, twqr_cev, twqr_aev, twqr_basis, twqr_reference "
        "FROM dbo.WQ_Indicators WHERE org_id = ? ORDER BY display_order",
        [org_id]
    )
    if not rows:
        return []
    return [
        {
            'id': r[0], 'code': r[1], 'name': r[2], 'unit': r[3],
            'lower_std': r[4], 'upper_std': r[5], 'reference': r[6], 'display_order': r[7],
            'risk_class': r[8],
            'twqr_lower': r[9], 'twqr_upper': r[10], 'twqr_cev': r[11], 'twqr_aev': r[12],
            'twqr_basis': r[13], 'twqr_reference': r[14],
        }
        for r in rows
    ]


def _fetch_hardness_bands(org_id):
    """{indicator_id: {hardness_class: (twqr_upper, cev, aev)}} for this org's
    hardness-dependent indicators (Cd, Cu, Pb)."""
    rows = _q(
        "SELECT b.indicator_id, b.hardness_class, b.twqr_upper, b.cev, b.aev "
        "FROM dbo.WQ_TWQR_HardnessBands b JOIN dbo.WQ_Indicators i ON i.indicator_id = b.indicator_id "
        "WHERE i.org_id = ?",
        [org_id]
    )
    result = {}
    for indicator_id, hardness_class, twqr_upper, cev, aev in (rows or []):
        result.setdefault(indicator_id, {})[hardness_class] = (twqr_upper, cev, aev)
    return result


def _season_for_date(reading_date):
    """Southern-Hemisphere wet/dry season, matching the CASE logic already
    used in the wq_schema_v2.sql backfill for existing readings."""
    if reading_date is None:
        return None
    month = reading_date.month if hasattr(reading_date, 'month') else None
    if month is None:
        return None
    return 'Wet' if month in (10, 11, 12, 1, 2, 3, 4) else 'Dry'


def _fetch_livestock_bands():
    """WQ_Livestock_TDS_Bands rows — universal reference table, not per-org."""
    rows = _q(
        "SELECT species, tds_min_mgL, tds_max_mgL, severity_rating, band_label "
        "FROM dbo.WQ_Livestock_TDS_Bands"
    )
    return [
        {'species': r[0], 'tds_min_mgL': r[1], 'tds_max_mgL': r[2],
         'severity_rating': r[3], 'band_label': r[4]}
        for r in (rows or [])
    ]


def _fetch_aquaculture_criteria():
    """WQ_Aquaculture_Criteria rows — universal reference table, not per-org."""
    rows = _q(
        "SELECT indicator_code, hardness_class, twqr_lower, twqr_upper, unit, note "
        "FROM dbo.WQ_Aquaculture_Criteria"
    )
    return [
        {'indicator_code': r[0], 'hardness_class': r[1], 'twqr_lower': r[2],
         'twqr_upper': r[3], 'unit': r[4], 'note': r[5]}
        for r in (rows or [])
    ]


def _fetch_station_context(org_id):
    """{station_id: {aquifer_type, depth_m}} for the org's stations — used
    only to optionally boost diagnostic-engine confidence; absent/NULL values
    simply mean that boost never fires."""
    rows = _q(
        "SELECT station_id, aquifer_type, depth_m FROM dbo.WQ_Stations WHERE org_id=?",
        [org_id]
    )
    return {r[0]: {'aquifer_type': r[1], 'depth_m': r[2]} for r in (rows or [])}


def _fetch_recent_indicator_history(station_id, indicator_code, before_date, limit=12):
    """Last `limit` measured values for one station+indicator, strictly before
    `before_date`, oldest-to-newest — feeds compute_temporal_stability() for
    the geogenic contamination signature."""
    rows = _q(
        f"""
        SELECT TOP {int(limit)} m.measured_val
        FROM dbo.WQ_Measurements m
        JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
        JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
        WHERE r.station_id = ? AND i.indicator_code = ? AND r.reading_date < ?
          AND m.measured_val IS NOT NULL
        ORDER BY r.reading_date DESC
        """,
        [station_id, indicator_code, before_date]
    )
    return [r[0] for r in reversed(rows or [])]


def _header_to_code_map(org_id):
    """Build dict: normalised header string → indicator_code for Excel parsing."""
    result = {}
    for ind in _fetch_indicators(org_id):
        name_lc = ind['name'].lower().strip()
        code     = ind['code']
        result[name_lc] = code
        if ind['unit']:
            result[f"{name_lc} ({ind['unit'].lower()})"] = code
        # also by code itself (uppercase and lowercase)
        result[code.lower()] = code
    return result


def _teal_fill():
    return PatternFill('solid', fgColor='0D9488')


def _header_font():
    return Font(color='FFFFFF', bold=True, size=10)


def _thin_border():
    s = Side(border_style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


# ── Dashboard ─────────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/dashboard', methods=['GET'])
@require_wq_auth()
def wq_dashboard():
    """Summary KPIs + latest trend data for the current organisation."""
    org = g.wq_org_id
    try:
        # Total active stations
        r_stations = _q(
            "SELECT COUNT(*) FROM dbo.WQ_Stations WHERE org_id=? AND is_active=1", [org]
        )
        total_stations = r_stations[0][0] if r_stations else 0

        # Total reading events
        r_readings = _q(
            "SELECT COUNT(*) FROM dbo.WQ_Readings WHERE org_id=?", [org]
        )
        total_readings = r_readings[0][0] if r_readings else 0

        # Latest reading date
        r_latest = _q(
            "SELECT MAX(r.reading_date) FROM dbo.WQ_Readings r WHERE r.org_id=?", [org]
        )
        latest_date = str(r_latest[0][0]) if r_latest and r_latest[0][0] else None

        # Compliance rate last 90 days
        r_comp = _q(
            """
            SELECT
                COUNT(CASE WHEN m.is_compliant = 1 THEN 1 END)  AS ok,
                COUNT(CASE WHEN m.is_compliant IS NOT NULL THEN 1 END) AS total
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings r ON r.reading_id = m.reading_id
            WHERE r.org_id = ?
              AND r.reading_date >= DATEADD(day, -90, CAST(GETDATE() AS DATE))
            """, [org]
        )
        if r_comp and r_comp[0][1]:
            compliance_rate = round(r_comp[0][0] / r_comp[0][1] * 100, 1)
            compliant_count = r_comp[0][0]
            total_meas      = r_comp[0][1]
        else:
            compliance_rate = None
            compliant_count = 0
            total_meas      = 0

        # Non-compliant flags last 30 days
        r_alerts = _q(
            """
            SELECT TOP 5
                s.station_name, i.indicator_name, i.unit, m.measured_val, m.flag,
                r.reading_date
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings    r ON r.reading_id    = m.reading_id
            JOIN dbo.WQ_Stations    s ON s.station_id    = r.station_id
            JOIN dbo.WQ_Indicators  i ON i.indicator_id  = m.indicator_id
            WHERE r.org_id = ?
              AND m.is_compliant = 0
              AND r.reading_date >= DATEADD(day, -30, CAST(GETDATE() AS DATE))
            ORDER BY r.reading_date DESC
            """, [org]
        )
        alerts = []
        if r_alerts:
            for row in r_alerts:
                alerts.append({
                    'station': row[0], 'indicator': row[1], 'unit': row[2],
                    'value': row[3], 'flag': row[4], 'date': str(row[5])
                })

        # EC trend (last 12 readings across all stations)
        r_trend = _q(
            """
            SELECT TOP 12 r.reading_date, AVG(m.measured_val) AS avg_ec
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.org_id = ? AND i.indicator_code = 'EC' AND m.measured_val IS NOT NULL
            GROUP BY r.reading_date
            ORDER BY r.reading_date DESC
            """, [org]
        )
        ec_trend = []
        if r_trend:
            for row in reversed(r_trend):
                ec_trend.append({'date': str(row[0]), 'value': round(row[1], 2)})

        return jsonify({
            'success': True,
            'kpis': {
                'total_stations': total_stations,
                'total_readings': total_readings,
                'latest_date': latest_date,
                'compliance_rate': compliance_rate,
                'compliant_count': compliant_count,
                'total_measured': total_meas,
            },
            'alerts': alerts,
            'ec_trend': ec_trend,
        })
    except Exception as exc:
        logger.exception('WQ dashboard error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Stations ──────────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/stations', methods=['GET'])
@require_wq_auth()
def list_stations():
    org = g.wq_org_id
    try:
        rows = _q(
            """
            SELECT s.station_id, s.station_code, s.station_name, s.description,
                   s.latitude, s.longitude, s.is_active, s.created_at,
                   COUNT(r.reading_id) AS reading_count,
                   s.aquifer_type, s.depth_m
            FROM dbo.WQ_Stations s
            LEFT JOIN dbo.WQ_Readings r ON r.station_id = s.station_id
            WHERE s.org_id = ?
            GROUP BY s.station_id, s.station_code, s.station_name, s.description,
                     s.latitude, s.longitude, s.is_active, s.created_at,
                     s.aquifer_type, s.depth_m
            ORDER BY s.station_name
            """, [org]
        )
        stations = []
        if rows:
            for r in rows:
                stations.append({
                    'id': r[0], 'code': r[1], 'name': r[2], 'description': r[3],
                    'lat': float(r[4]) if r[4] is not None else None,
                    'lng': float(r[5]) if r[5] is not None else None,
                    'active': bool(r[6]), 'created_at': str(r[7]),
                    'reading_count': r[8],
                    'aquifer_type': r[9],
                    'depth_m': float(r[10]) if r[10] is not None else None,
                })
        return jsonify({'success': True, 'stations': stations})
    except Exception as exc:
        logger.exception('WQ list_stations error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/stations', methods=['POST'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def create_station():
    org = g.wq_org_id
    data = request.get_json(force=True) or {}
    code = (data.get('code') or '').strip()
    name = (data.get('name') or '').strip()
    if not code or not name:
        return jsonify({'success': False, 'error': 'Station code and name are required.'}), 400
    aquifer_type = (data.get('aquifer_type') or '').strip() or None
    if aquifer_type and aquifer_type not in ('unconfined', 'semi-confined', 'confined'):
        return jsonify({'success': False, 'error': "aquifer_type must be one of: unconfined, semi-confined, confined."}), 400
    try:
        _q(
            """
            INSERT INTO dbo.WQ_Stations (station_code, station_name, description, latitude, longitude, org_id, aquifer_type, depth_m)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [code, name, data.get('description'), _safe_float(data.get('lat')),
             _safe_float(data.get('lng')), org, aquifer_type, _safe_float(data.get('depth_m'))],
            fetch=False, commit=True
        )
        new_row = _q(
            "SELECT TOP 1 station_id FROM dbo.WQ_Stations WHERE station_code=? AND org_id=? ORDER BY station_id DESC",
            [code, org]
        )
        new_id = new_row[0][0] if new_row else None
        return jsonify({'success': True, 'station_id': new_id}), 201
    except Exception as exc:
        msg = str(exc)
        if '2601' in msg or '2627' in msg:
            return jsonify({'success': False, 'error': f"Station code '{code}' already exists."}), 409
        logger.exception('WQ create_station error')
        return jsonify({'success': False, 'error': msg}), 500


@wq_bp.route('/api/wq/stations/<int:station_id>', methods=['PUT'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def update_station(station_id):
    org = g.wq_org_id
    data = request.get_json(force=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Station name is required.'}), 400
    aquifer_type = (data.get('aquifer_type') or '').strip() or None
    if aquifer_type and aquifer_type not in ('unconfined', 'semi-confined', 'confined'):
        return jsonify({'success': False, 'error': "aquifer_type must be one of: unconfined, semi-confined, confined."}), 400
    try:
        _q(
            """
            UPDATE dbo.WQ_Stations
            SET station_name=?, description=?, latitude=?, longitude=?, is_active=?, aquifer_type=?, depth_m=?
            WHERE station_id=? AND org_id=?
            """,
            [name, data.get('description'), _safe_float(data.get('lat')),
             _safe_float(data.get('lng')), 1 if data.get('active', True) else 0,
             aquifer_type, _safe_float(data.get('depth_m')),
             station_id, org],
            fetch=False, commit=True
        )
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ update_station error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/stations/<int:station_id>', methods=['DELETE'])
@require_wq_auth(roles=['admin', 'standard_user'])
def delete_station(station_id):
    org = g.wq_org_id
    try:
        _q(
            "DELETE FROM dbo.WQ_Stations WHERE station_id=? AND org_id=?",
            [station_id, org], fetch=False, commit=True
        )
        return jsonify({'success': True})
    except Exception as exc:
        msg = str(exc)
        if 'FK_WQR_Station' in msg or 'REFERENCE' in msg.upper():
            return jsonify({'success': False, 'error': 'Station has readings attached. Delete readings first.'}), 409
        logger.exception('WQ delete_station error')
        return jsonify({'success': False, 'error': msg}), 500


# ── Indicators ────────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/indicators', methods=['GET'])
@require_wq_auth()
def list_indicators():
    try:
        return jsonify({'success': True, 'indicators': _fetch_indicators(g.wq_org_id)})
    except Exception as exc:
        logger.exception('WQ list_indicators error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Readings list + delete ────────────────────────────────────────────────────

@wq_bp.route('/api/wq/readings', methods=['GET'])
@require_wq_auth()
def list_readings():
    org         = g.wq_org_id
    station_id  = request.args.get('station_id', type=int)
    date_from   = request.args.get('date_from')
    date_to     = request.args.get('date_to')
    page        = request.args.get('page', 1, type=int)
    per_page    = min(request.args.get('per_page', 20, type=int), 100)

    where_clauses = ['r.org_id = ?']
    params        = [org]
    if station_id:
        where_clauses.append('r.station_id = ?')
        params.append(station_id)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)

    where_sql = ' AND '.join(where_clauses)
    offset    = (page - 1) * per_page

    try:
        count_row = _q(
            f"SELECT COUNT(*) FROM dbo.WQ_Readings r WHERE {where_sql}", params
        )
        total = count_row[0][0] if count_row else 0

        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, s.station_name, s.station_code,
                   r.source_file, r.created_at,
                   COUNT(m.meas_id) AS meas_count,
                   SUM(CASE WHEN m.is_compliant = 0 THEN 1 ELSE 0 END) AS fails
            FROM dbo.WQ_Readings r
            JOIN dbo.WQ_Stations s ON s.station_id = r.station_id
            LEFT JOIN dbo.WQ_Measurements m ON m.reading_id = r.reading_id
            WHERE {where_sql}
            GROUP BY r.reading_id, r.reading_date, s.station_name, s.station_code,
                     r.source_file, r.created_at
            ORDER BY r.reading_date DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
            """,
            params + [offset, per_page]
        )
        readings = []
        if rows:
            for row in rows:
                readings.append({
                    'id': row[0], 'date': str(row[1]), 'station': row[2],
                    'station_code': row[3], 'source_file': row[4],
                    'created_at': str(row[5]), 'meas_count': row[6], 'fails': row[7] or 0
                })
        return jsonify({
            'success': True, 'readings': readings,
            'total': total, 'page': page, 'per_page': per_page
        })
    except Exception as exc:
        logger.exception('WQ list_readings error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/readings/<int:reading_id>', methods=['DELETE'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def delete_reading(reading_id):
    org = g.wq_org_id
    try:
        check = _q(
            "SELECT reading_id FROM dbo.WQ_Readings WHERE reading_id=? AND org_id=?",
            [reading_id, org]
        )
        if not check:
            return jsonify({'success': False, 'error': 'Reading not found.'}), 404
        _q("DELETE FROM dbo.WQ_Readings WHERE reading_id=?", [reading_id], fetch=False, commit=True)
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ delete_reading error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Reading detail (all measurements for one event) ───────────────────────────

@wq_bp.route('/api/wq/readings/<int:reading_id>/measurements', methods=['GET'])
@require_wq_auth()
def reading_measurements(reading_id):
    org = g.wq_org_id
    try:
        check = _q(
            "SELECT reading_id FROM dbo.WQ_Readings WHERE reading_id=? AND org_id=?",
            [reading_id, org]
        )
        if not check:
            return jsonify({'success': False, 'error': 'Reading not found.'}), 404
        rows = _q(
            """
            SELECT i.indicator_code, i.indicator_name, i.unit, i.lower_std, i.upper_std,
                   m.measured_val, m.is_compliant, m.flag
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE m.reading_id = ?
            ORDER BY i.display_order
            """, [reading_id]
        )
        measurements = []
        if rows:
            for r in rows:
                measurements.append({
                    'code': r[0], 'name': r[1], 'unit': r[2],
                    'lower_std': r[3], 'upper_std': r[4],
                    'value': r[5], 'compliant': bool(r[6]) if r[6] is not None else None,
                    'flag': r[7],
                })
        return jsonify({'success': True, 'measurements': measurements})
    except Exception as exc:
        logger.exception('WQ reading_measurements error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Compliance matrix ─────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/compliance', methods=['GET'])
@require_wq_auth()
def compliance_matrix():
    """
    Returns a 2-D compliance matrix: rows = indicators, cols = recent reading events.
    Each cell: {value, compliant, flag}.  Used for the heatmap visualisation.

    ?standard=sans241 (default) — SANS 241:2015 drinking-water standard.
    ?standard=twqr             — DWAF Vol 7 (1996) Target Water Quality Range
                                  for aquatic ecosystems/resource water. Rows
                                  whose twqr_basis is 'not_applicable' (or NULL)
                                  are marked not_assessed=True on every cell,
                                  since TWQR has no fixed value for them (see
                                  twqr_standard_migration.sql).
    ?date_from / ?date_to      — optional; when omitted, shows the true latest
                                  N readings (default "always current" status
                                  view). When set, restricts to readings in
                                  that window, matching /api/wq/trend and
                                  /api/wq/readings on the same filter bar.
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    limit      = min(request.args.get('limit', 8, type=int), 20)
    standard   = request.args.get('standard', 'sans241').lower()
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')
    if standard not in ('sans241', 'twqr'):
        standard = 'sans241'
    compliant_col = 'm.is_compliant_twqr' if standard == 'twqr' else 'm.is_compliant'
    flag_col      = 'm.flag_twqr' if standard == 'twqr' else 'm.flag'

    where_clauses = ['r.org_id = ?']
    params        = [org]
    if station_id:
        where_clauses.append('r.station_id = ?')
        params.append(station_id)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)
    where_sql = ' AND '.join(where_clauses)

    try:
        date_rows = _q(
            f"""
            SELECT TOP {limit} r.reading_id, r.reading_date, s.station_name
            FROM dbo.WQ_Readings r
            JOIN dbo.WQ_Stations s ON s.station_id = r.station_id
            WHERE {where_sql}
            ORDER BY r.reading_date DESC
            """, params
        )
        if not date_rows:
            return jsonify({'success': True, 'cols': [], 'rows': [], 'standard': standard})

        reading_ids  = [str(row[0]) for row in date_rows]
        col_labels   = [{'id': row[0], 'date': str(row[1]), 'station': row[2]} for row in date_rows]

        meas_rows = _q(
            f"""
            SELECT m.reading_id, i.indicator_code, i.indicator_name, i.unit,
                   m.measured_val, {compliant_col}, {flag_col}, i.twqr_basis
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Indicators   i ON i.indicator_id  = m.indicator_id
            JOIN dbo.WQ_Readings     r ON r.reading_id    = m.reading_id
            WHERE m.reading_id IN ({','.join(['?']*len(reading_ids))})
            ORDER BY i.display_order
            """, reading_ids
        )

        # Build dict: indicator_code → {reading_id → cell}
        ind_map  = {}
        ind_meta = {}
        if meas_rows:
            for mr in meas_rows:
                rid, code, name, unit, val, comp, flag, twqr_basis = mr
                not_assessed = standard == 'twqr' and twqr_basis in (None, 'not_applicable')
                if code not in ind_map:
                    ind_map[code]  = {}
                    ind_meta[code] = {'code': code, 'name': name, 'unit': unit, 'not_assessed': not_assessed}
                ind_map[code][rid] = {
                    'value': round(val, 4) if val is not None else None,
                    'compliant': bool(comp) if comp is not None else None,
                    'flag': flag,
                    'not_assessed': not_assessed,
                }

        matrix_rows = []
        for code, cells in ind_map.items():
            row_cells = [cells.get(int(col['id'])) for col in col_labels]
            matrix_rows.append({
                'indicator': ind_meta[code],
                'cells': row_cells,
            })

        return jsonify({'success': True, 'cols': col_labels, 'rows': matrix_rows, 'standard': standard})
    except Exception as exc:
        logger.exception('WQ compliance_matrix error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Trend series ──────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/trend', methods=['GET'])
@require_wq_auth()
def wq_trend():
    """Time-series of a single indicator for one (or all) stations."""
    org          = g.wq_org_id
    indicator_code = request.args.get('indicator', 'EC')
    station_id   = request.args.get('station_id', type=int)
    date_from    = request.args.get('date_from')
    date_to      = request.args.get('date_to')

    where_clauses = ['r.org_id = ?', 'i.indicator_code = ?']
    params        = [org, indicator_code]
    if station_id:
        where_clauses.append('r.station_id = ?')
        params.append(station_id)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)
    where_sql = ' AND '.join(where_clauses)

    try:
        rows = _q(
            f"""
            SELECT r.reading_date, s.station_name, m.measured_val, m.is_compliant, m.flag,
                   i.upper_std, i.lower_std, i.unit
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id    = m.reading_id
            JOIN dbo.WQ_Stations   s ON s.station_id    = r.station_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id  = m.indicator_id
            WHERE {where_sql} AND m.measured_val IS NOT NULL
            ORDER BY r.reading_date ASC
            """, params
        )

        series = []
        upper_std = lower_std = unit = None
        if rows:
            for row in rows:
                if upper_std is None:
                    upper_std = row[5]
                    lower_std = row[6]
                    unit      = row[7]
                series.append({
                    'date':      str(row[0]),
                    'station':   row[1],
                    'value':     round(row[2], 4),
                    'compliant': bool(row[3]) if row[3] is not None else None,
                    'flag':      row[4],
                })

        return jsonify({
            'success': True,
            'indicator_code': indicator_code,
            'unit': unit,
            'upper_std': upper_std,
            'lower_std': lower_std,
            'series': series,
        })
    except Exception as exc:
        logger.exception('WQ trend error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Excel upload ──────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/upload', methods=['POST'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def wq_upload():
    """
    Accept an Excel file (.xlsx).  Expected format (generated by /api/wq/template):
      Row 1 : headers — 'Reading Date', 'Station Name', <indicator names>
      Row 2+: data rows
    Each row becomes one WQ_Readings record + measurements per indicator column found.
    """
    org     = g.wq_org_id
    user_id = g.wq_user_id

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file attached.'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Only .xlsx / .xls files are accepted.'}), 400

    try:
        df = pd.read_excel(f, header=0, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        # Map column index → indicator
        header_map = _header_to_code_map(org)
        col_to_code = {}
        for col in df.columns:
            norm = col.lower().strip()
            if norm in header_map:
                col_to_code[col] = header_map[norm]

        if not col_to_code:
            return jsonify({'success': False, 'error': 'No recognisable indicator columns found. Please use the provided template.'}), 400

        # Pre-fetch org stations → name lookup
        st_rows = _q(
            "SELECT station_name, station_id FROM dbo.WQ_Stations WHERE org_id=? AND is_active=1",
            [org]
        )
        station_lookup = {r[0].strip().lower(): r[1] for r in st_rows} if st_rows else {}

        # Pre-fetch indicators → thresholds + metadata for alert dispatch.
        # Must be org-scoped: WQ_Indicators is tenant-owned, so an unfiltered
        # query here would let two orgs' rows collide in this dict (keyed by
        # indicator_code, which is only unique per-org now) and silently write
        # measurements against the wrong tenant's indicator_id.
        ind_lookup = {
            ind['code']: {
                'id': ind['id'], 'lower': ind['lower_std'], 'upper': ind['upper_std'],
                'name': ind['name'], 'unit': ind['unit'], 'risk_class': ind['risk_class'],
                'twqr_lower': ind['twqr_lower'], 'twqr_upper': ind['twqr_upper'],
                'twqr_basis': ind['twqr_basis'],
            }
            for ind in _fetch_indicators(org)
        }
        hardness_bands = _fetch_hardness_bands(org)

        # Pre-fetched once per upload (not per row) for the diagnostic engine.
        station_context = _fetch_station_context(org)
        stability_cache  = {}  # {(station_id, indicator_code): compute_temporal_stability() result}

        inserted  = 0
        skipped   = 0
        errors    = []

        for row_i, row in df.iterrows():
            human_row = row_i + 2  # Excel row number

            # Reading date
            raw_date = str(row.get('Reading Date', '') or '').strip()
            if not raw_date or raw_date.lower() in ('nan', 'reading date'):
                continue
            try:
                reading_date = pd.to_datetime(raw_date, dayfirst=False).date()
            except Exception:
                errors.append(f"Row {human_row}: unrecognised date '{raw_date}'")
                skipped += 1
                continue

            # Station
            raw_station = str(row.get('Station Name', '') or '').strip()
            if not raw_station or raw_station.lower() == 'nan':
                errors.append(f"Row {human_row}: station name is blank.")
                skipped += 1
                continue
            station_id = station_lookup.get(raw_station.lower())
            if not station_id:
                errors.append(f"Row {human_row}: station '{raw_station}' not found.")
                skipped += 1
                continue

            # Insert or find reading event
            existing = _q(
                "SELECT reading_id FROM dbo.WQ_Readings WHERE station_id=? AND reading_date=?",
                [station_id, reading_date]
            )
            if existing:
                reading_id = existing[0][0]
            else:
                _q(
                    "INSERT INTO dbo.WQ_Readings (station_id, reading_date, source_file, uploaded_by, org_id, season) VALUES (?,?,?,?,?,?)",
                    [station_id, reading_date, f.filename, user_id, org, _season_for_date(reading_date)],
                    fetch=False, commit=True
                )
                new_r = _q(
                    "SELECT TOP 1 reading_id FROM dbo.WQ_Readings WHERE station_id=? AND reading_date=?",
                    [station_id, reading_date]
                )
                reading_id = new_r[0][0]

            # This row's Total Hardness value (if present), needed to pick the
            # correct DWAF TWQR band for the hardness-dependent metals (Cd/Cu/Pb).
            hardness_val = None
            for col, code in col_to_code.items():
                if code == 'HARD':
                    hardness_val = _safe_float(row.get(col))
                    break

            # Measurements
            measured_by_code = {}
            for col, code in col_to_code.items():
                raw_val = _safe_float(row.get(col))
                measured_by_code[code] = raw_val
                ind     = ind_lookup.get(code)
                if not ind:
                    continue
                compliant, flag = _compute_compliance(raw_val, ind['lower'], ind['upper'])
                compliant_twqr, flag_twqr = _compute_twqr_compliance(raw_val, ind, hardness_val, hardness_bands)

                ex_m = _q(
                    "SELECT meas_id FROM dbo.WQ_Measurements m JOIN dbo.WQ_Indicators i ON i.indicator_id=m.indicator_id "
                    "WHERE m.reading_id=? AND i.indicator_code=?",
                    [reading_id, code]
                )
                if ex_m:
                    _q(
                        "UPDATE dbo.WQ_Measurements SET measured_val=?, is_compliant=?, flag=?, "
                        "is_compliant_twqr=?, flag_twqr=?, quality_flag='Approved' "
                        "WHERE meas_id=?",
                        [raw_val, compliant, flag, compliant_twqr, flag_twqr, ex_m[0][0]],
                        fetch=False, commit=True
                    )
                else:
                    _q(
                        "INSERT INTO dbo.WQ_Measurements "
                        "(reading_id, indicator_id, measured_val, is_compliant, flag, "
                        " is_compliant_twqr, flag_twqr, quality_flag) "
                        "VALUES (?,?,?,?,?,?,?,'Approved')",
                        [reading_id, ind['id'], raw_val, compliant, flag, compliant_twqr, flag_twqr],
                        fetch=False, commit=True
                    )

                # Dispatch alerts for health-risk non-compliant measurements
                if compliant is False and ind.get('risk_class') == 'health' and flag:
                    threshold = ind['upper'] if flag == 'ABOVE' else ind['lower']
                    try:
                        dispatch_alert(
                            db_query_fn    = _q,
                            org_id         = org,
                            reading_id     = reading_id,
                            indicator_id   = ind['id'],
                            indicator_code = code,
                            indicator_name = ind['name'],
                            unit           = ind['unit'] or '',
                            measured_val   = raw_val,
                            threshold_value= threshold,
                            threshold_type = flag,
                            station_name   = raw_station,
                            reading_date   = str(reading_date),
                        )
                    except Exception:
                        logger.exception('Alert dispatch failed for %s reading_id=%s', code, reading_id)

                    # Auto-capture a draft corrective action so the violation is
                    # tracked to closure even if nobody acts on the alert email.
                    # Guarded against re-uploads of the same reading creating
                    # duplicate entries: only insert if there's no existing
                    # auto-generated action for this exact reading+indicator.
                    try:
                        already = _q(
                            "SELECT 1 FROM dbo.WQ_Corrective_Actions "
                            "WHERE reading_id=? AND indicator_id=? AND auto_generated=1",
                            [reading_id, ind['id']]
                        )
                        if not already:
                            _q(
                                """
                                INSERT INTO dbo.WQ_Corrective_Actions
                                    (org_id, station_id, indicator_id, reading_id, description, auto_generated)
                                VALUES (?,?,?,?,?,1)
                                """,
                                [org, station_id, ind['id'], reading_id,
                                 f"Auto-generated: {ind['name']} measured {raw_val}{ind['unit'] or ''} "
                                 f"({flag} threshold {threshold}) at {raw_station} on {reading_date}."],
                                fetch=False, commit=True
                            )
                    except Exception:
                        logger.exception('Auto corrective-action capture failed for %s reading_id=%s', code, reading_id)

            # Contamination source-attribution engine — runs once per reading,
            # after the full measurement set for this row is known (it needs
            # multiple indicators simultaneously, e.g. pH+SO4+Fe for AMD, so it
            # cannot run per-column like the block above). Isolated in its own
            # try/except so a diagnostic-engine failure never breaks SANS/TWQR
            # ingestion, matching the fault-isolation style already used for
            # alert dispatch and corrective-action auto-capture above.
            try:
                ctx_station = station_context.get(station_id, {})
                historical_stability = {}
                for stab_code in ('FE', 'MN', 'FL', 'EC'):
                    cache_key = (station_id, stab_code)
                    if cache_key not in stability_cache:
                        history = _fetch_recent_indicator_history(station_id, stab_code, reading_date)
                        stability_cache[cache_key] = compute_temporal_stability(history)
                    historical_stability[stab_code] = stability_cache[cache_key]

                diag_context = {
                    'season':        _season_for_date(reading_date),
                    'aquifer_type':  ctx_station.get('aquifer_type'),
                    'depth_m':       float(ctx_station['depth_m']) if ctx_station.get('depth_m') is not None else None,
                    'historical_stability': historical_stability,
                }
                diag = diagnose_contamination_source(measured_by_code, diag_context)
                candidates = [c for c in diag['candidates'] if c['confidence_pct'] >= 15]

                # Idempotent re-upload: replace this reading's findings wholesale.
                _q("DELETE FROM dbo.WQ_Diagnostic_Findings WHERE reading_id=?", [reading_id], fetch=False, commit=True)

                for c in candidates:
                    explanation = _ascii_safe(explain_diagnostic_finding(c))
                    matched_json = _ascii_safe(json.dumps(c['matched_signals']))
                    contra_json  = _ascii_safe(json.dumps(c['contradicting_signals']))
                    modifiers_json = _ascii_safe(json.dumps(c['contextual_modifiers']))
                    _q(
                        """
                        INSERT INTO dbo.WQ_Diagnostic_Findings
                            (reading_id, station_id, org_id, source_code, source_label, confidence_pct,
                             matched_signals, contradicting_signals, contextual_modifiers_applied, explanation_text)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        [reading_id, station_id, org, c['source_code'], _ascii_safe(c['source_label']), c['confidence_pct'],
                         matched_json, contra_json, modifiers_json, explanation],
                        fetch=False, commit=True
                    )

                    # High-confidence AMD/sewage findings auto-create a draft
                    # Risk Register entry (mirrors the corrective-action
                    # auto-capture above). Guarded against duplicate spam on
                    # re-upload via a hazard-description existence check.
                    if c['confidence_pct'] >= 70 and c['source_code'] in ('AMD', 'SEWAGE'):
                        already_risk = _q(
                            "SELECT 1 FROM dbo.WQ_Risk_Register WHERE station_id=? AND hazard_description LIKE ?",
                            [station_id, f"%{c['source_label']}%reading on {reading_date}%"]
                        )
                        if not already_risk:
                            _q(
                                """
                                INSERT INTO dbo.WQ_Risk_Register
                                    (org_id, station_id, hazard_description, hazard_source, likelihood, severity)
                                VALUES (?,?,?,?,?,?)
                                """,
                                [org, station_id,
                                 f"Auto-suggested: {c['source_label']} suspected ({c['confidence_pct']}% confidence) "
                                 f"at {raw_station} for reading on {reading_date}.",
                                 'Diagnostic engine', 3, 4],
                                fetch=False, commit=True
                            )
            except Exception:
                logger.exception('Diagnostic engine failed for reading_id=%s', reading_id)

            inserted += 1

        return jsonify({
            'success': True,
            'inserted': inserted,
            'skipped': skipped,
            'errors': errors[:20],
        })
    except Exception as exc:
        logger.exception('WQ upload error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Upload template download ──────────────────────────────────────────────────

@wq_bp.route('/api/wq/template', methods=['GET'])
@require_wq_auth()
def wq_template():
    """Return a blank Excel upload template with correct headers."""
    try:
        indicators = _fetch_indicators(g.wq_org_id)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Water Quality Data'

        header_fill = _teal_fill()
        header_font = _header_font()

        headers = ['Reading Date', 'Station Name']
        for ind in indicators:
            label = ind['name'] + (f" ({ind['unit']})" if ind['unit'] else '')
            headers.append(label)

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 16)

        ws.row_dimensions[1].height = 28

        # Sample row
        ws.cell(row=2, column=1, value='2024-01-15')
        ws.cell(row=2, column=2, value='Station A')
        for ci in range(3, len(headers) + 1):
            ws.cell(row=2, column=ci, value='')

        # Notes row
        note_cell = ws.cell(
            row=3, column=1,
            value='NOTE: Date format YYYY-MM-DD. Station Name must match an existing station exactly.'
        )
        note_cell.font = Font(italic=True, color='888888', size=9)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='water_quality_upload_template.xlsx',
        )
    except Exception as exc:
        logger.exception('WQ template error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Data export ───────────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/export', methods=['GET'])
@require_wq_auth()
def wq_export():
    """Export all readings (filtered by station/date) as a flat Excel sheet."""
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    where_clauses = ['r.org_id = ?']
    params        = [org]
    if station_id:
        where_clauses.append('r.station_id = ?')
        params.append(station_id)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)
    where_sql = ' AND '.join(where_clauses)

    try:
        rows = _q(
            f"""
            SELECT r.reading_date, s.station_name, s.station_code,
                   i.indicator_code, i.indicator_name, i.unit,
                   m.measured_val, m.is_compliant, m.flag
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Stations   s ON s.station_id   = r.station_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE {where_sql}
            ORDER BY r.reading_date DESC, s.station_name, i.display_order
            """, params
        )

        indicators = _fetch_indicators(org)
        ind_codes  = [i['code'] for i in indicators]

        wb  = Workbook()
        ws  = wb.active
        ws.title = 'WQ Export'

        fill = _teal_fill()
        hdr_font = _header_font()

        fixed_headers = ['Reading Date', 'Station Name', 'Station Code']
        all_headers   = fixed_headers + [
            i['name'] + (f" ({i['unit']})" if i['unit'] else '') for i in indicators
        ] + ['Compliance Status']

        for ci, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill   = fill
            cell.font   = hdr_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 2, 14)

        # Pivot: group by (date, station)
        from collections import defaultdict
        pivot = defaultdict(dict)
        pivot_meta = {}
        if rows:
            for row in rows:
                date_, sname, scode, icode, _, _, val, comp, flag = row
                key = (str(date_), sname, scode)
                pivot[key][icode] = (val, bool(comp) if comp is not None else None)
                pivot_meta[key] = True

        row_i = 2
        for key in sorted(pivot.keys()):
            date_, sname, scode = key
            ws.cell(row=row_i, column=1, value=date_)
            ws.cell(row=row_i, column=2, value=sname)
            ws.cell(row=row_i, column=3, value=scode)
            cell_data  = pivot[key]
            all_comp   = [v[1] for v in cell_data.values() if v[1] is not None]
            overall    = 'Compliant' if all_comp and all(all_comp) else ('Non-Compliant' if any(c is False for c in all_comp) else 'No Standard')
            for ci, ind in enumerate(indicators, start=4):
                pair = cell_data.get(ind['code'])
                ws.cell(row=row_i, column=ci, value=pair[0] if pair else None)
            ws.cell(row=row_i, column=len(fixed_headers) + len(indicators) + 1, value=overall)
            row_i += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'wq_export_{date.today().isoformat()}.xlsx',
        )
    except Exception as exc:
        logger.exception('WQ export error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Quality flag bulk PATCH ───────────────────────────────────────────────────

@wq_bp.route('/api/wq/measurements/quality-flags', methods=['PATCH'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def patch_quality_flags():
    """
    Bulk-update quality_flag on a set of measurement IDs.
    Body: {meas_ids: [int], quality_flag: str}
    Valid flags: Approved | Preliminary | Estimated | Questionable | Rejected
    """
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}
    meas_ids = data.get('meas_ids', [])
    new_flag = (data.get('quality_flag') or '').strip()

    VALID_FLAGS = {'Approved', 'Preliminary', 'Estimated', 'Questionable', 'Rejected'}
    if new_flag not in VALID_FLAGS:
        return jsonify({'success': False, 'error': f"Invalid flag. Must be one of: {', '.join(VALID_FLAGS)}"}), 400
    if not meas_ids or not isinstance(meas_ids, list):
        return jsonify({'success': False, 'error': 'meas_ids must be a non-empty list.'}), 400

    try:
        placeholders = ','.join(['?'] * len(meas_ids))
        _q(
            f"""
            UPDATE dbo.WQ_Measurements
            SET quality_flag = ?
            WHERE meas_id IN ({placeholders})
              AND reading_id IN (SELECT reading_id FROM dbo.WQ_Readings WHERE org_id = ?)
            """,
            [new_flag] + list(meas_ids) + [org],
            fetch=False, commit=True
        )
        return jsonify({'success': True, 'updated': len(meas_ids)})
    except Exception as exc:
        logger.exception('WQ patch_quality_flags error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── CCME Water Quality Index ──────────────────────────────────────────────────

@wq_bp.route('/api/wq/wqi', methods=['GET'])
@require_wq_auth()
def wq_wqi():
    """
    Compute CCME-WQI for a station over a date range.
    Query: station_id (required), date_from, date_to, approved_only (1|0, default 1)
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')
    approved_only = request.args.get('approved_only', '1') != '0'

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    try:
        # Verify station belongs to org
        st_check = _q("SELECT station_id FROM dbo.WQ_Stations WHERE station_id=? AND org_id=?", [station_id, org])
        if not st_check:
            return jsonify({'success': False, 'error': 'Station not found.'}), 404

        where_extra = ''
        params = [station_id]
        if date_from:
            where_extra += ' AND r.reading_date >= ?'
            params.append(date_from)
        if date_to:
            where_extra += ' AND r.reading_date <= ?'
            params.append(date_to)

        rows = _q(
            f"""
            SELECT i.indicator_code, m.measured_val, i.lower_std, i.upper_std,
                   m.quality_flag, m.is_censored
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? {where_extra}
            """,
            params
        )

        measurements = []
        if rows:
            for row in rows:
                measurements.append({
                    'indicator_code': row[0],
                    'measured_val':   row[1],
                    'lower_std':      row[2],
                    'upper_std':      row[3],
                    'quality_flag':   row[4],
                    'is_censored':    bool(row[5]) if row[5] is not None else False,
                })

        result = compute_ccme_wqi(measurements, approved_only=approved_only)
        return jsonify({'success': True, 'wqi': result})
    except Exception as exc:
        logger.exception('WQ wqi error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Ion Balance Error ─────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/readings/<int:reading_id>/ion-balance', methods=['GET'])
@require_wq_auth()
def reading_ion_balance(reading_id):
    """Compute charge balance error for one reading event."""
    org = g.wq_org_id
    try:
        check = _q("SELECT reading_id FROM dbo.WQ_Readings WHERE reading_id=? AND org_id=?", [reading_id, org])
        if not check:
            return jsonify({'success': False, 'error': 'Reading not found.'}), 404

        rows = _q(
            """
            SELECT i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE m.reading_id = ? AND m.measured_val IS NOT NULL AND m.is_censored = 0
            """,
            [reading_id]
        )
        conc = {r[0]: r[1] for r in rows} if rows else {}
        result = compute_ion_balance_error(conc)

        # Cache IBE in WQ_Readings
        if result['IBE_pct'] is not None:
            _q(
                "UPDATE dbo.WQ_Readings SET ion_balance_error_pct=? WHERE reading_id=?",
                [result['IBE_pct'], reading_id],
                fetch=False, commit=True
            )

        return jsonify({'success': True, 'ion_balance': result})
    except Exception as exc:
        logger.exception('WQ ion_balance error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Blue Drop Scoring ─────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/blue-drop', methods=['GET'])
@require_wq_auth()
def wq_blue_drop():
    """
    Compute Blue Drop compliance score for the organisation.
    Query: date_from, date_to  (optional)
    """
    org       = g.wq_org_id
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')

    where_extra = ''
    params = [org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT i.blue_drop_category,
                   COUNT(*) AS total_tests,
                   SUM(CASE WHEN m.is_compliant = 1 THEN 1 ELSE 0 END) AS passed
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.org_id = ?
              AND i.blue_drop_category IS NOT NULL
              AND m.is_compliant IS NOT NULL
              AND m.quality_flag = 'Approved'
              {where_extra}
            GROUP BY i.blue_drop_category
            """,
            params
        )

        compliance_by_category = {}
        category_counts = {}
        if rows:
            for row in rows:
                cat, total, passed = row
                if cat and total > 0:
                    compliance_by_category[cat] = round(passed / total * 100.0, 1)
                    category_counts[cat] = {'total': total, 'passed': passed}

        result = compute_blue_drop_score(compliance_by_category)
        result['category_counts'] = category_counts
        return jsonify({'success': True, 'blue_drop': result})
    except Exception as exc:
        logger.exception('WQ blue_drop error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Alert Config CRUD ─────────────────────────────────────────────────────────

@wq_bp.route('/api/wq/alerts/config', methods=['GET'])
@require_wq_auth()
def list_alert_configs():
    org = g.wq_org_id
    try:
        rows = _q(
            """
            SELECT ac.alert_id, i.indicator_name, i.unit, s.station_name,
                   ac.threshold_type, ac.threshold_value, ac.alert_email,
                   ac.alert_whatsapp, ac.is_active, ac.created_at
            FROM dbo.WQ_Alert_Config ac
            JOIN dbo.WQ_Indicators i ON i.indicator_id = ac.indicator_id
            LEFT JOIN dbo.WQ_Stations s ON s.station_id = ac.station_id
            WHERE ac.org_id = ?
            ORDER BY ac.created_at DESC
            """,
            [org]
        )
        configs = []
        if rows:
            for r in rows:
                configs.append({
                    'id': r[0], 'indicator': r[1], 'unit': r[2],
                    'station': r[3] or 'All stations',
                    'threshold_type': r[4], 'threshold_value': r[5],
                    'alert_email': r[6], 'alert_whatsapp': r[7],
                    'is_active': bool(r[8]), 'created_at': str(r[9]),
                })
        return jsonify({'success': True, 'configs': configs})
    except Exception as exc:
        logger.exception('WQ list_alert_configs error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/alerts/config', methods=['POST'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def create_alert_config():
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}

    indicator_id    = data.get('indicator_id')
    station_id      = data.get('station_id') or None
    threshold_type  = (data.get('threshold_type') or '').strip().upper()
    threshold_value = _safe_float(data.get('threshold_value'))
    alert_email     = (data.get('alert_email') or '').strip() or None
    alert_whatsapp  = (data.get('alert_whatsapp') or '').strip() or None

    if not indicator_id:
        return jsonify({'success': False, 'error': 'indicator_id is required.'}), 400
    if threshold_type not in ('ABOVE', 'BELOW', 'HEALTH'):
        return jsonify({'success': False, 'error': "threshold_type must be ABOVE, BELOW, or HEALTH."}), 400
    if threshold_value is None:
        return jsonify({'success': False, 'error': 'threshold_value is required.'}), 400
    if not alert_email and not alert_whatsapp:
        return jsonify({'success': False, 'error': 'At least one of alert_email or alert_whatsapp is required.'}), 400

    try:
        _q(
            """
            INSERT INTO dbo.WQ_Alert_Config
                (org_id, indicator_id, station_id, threshold_type, threshold_value,
                 alert_email, alert_whatsapp)
            VALUES (?,?,?,?,?,?,?)
            """,
            [org, indicator_id, station_id, threshold_type, threshold_value,
             alert_email, alert_whatsapp],
            fetch=False, commit=True
        )
        new_row = _q(
            "SELECT TOP 1 alert_id FROM dbo.WQ_Alert_Config WHERE org_id=? ORDER BY alert_id DESC",
            [org]
        )
        return jsonify({'success': True, 'alert_id': new_row[0][0] if new_row else None}), 201
    except Exception as exc:
        logger.exception('WQ create_alert_config error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/alerts/config/<int:alert_id>', methods=['DELETE'])
@require_wq_auth(roles=['admin', 'standard_user'])
def delete_alert_config(alert_id):
    org = g.wq_org_id
    try:
        _q(
            "DELETE FROM dbo.WQ_Alert_Config WHERE alert_id=? AND org_id=?",
            [alert_id, org], fetch=False, commit=True
        )
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ delete_alert_config error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/alerts/log', methods=['GET'])
@require_wq_auth()
def alert_log():
    org      = g.wq_org_id
    page     = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    offset   = (page - 1) * per_page
    try:
        total_row = _q("SELECT COUNT(*) FROM dbo.WQ_Alert_Log WHERE org_id=?", [org])
        total     = total_row[0][0] if total_row else 0

        rows = _q(
            """
            SELECT al.log_id, i.indicator_name, al.measured_val,
                   al.channel, al.recipient, al.status, al.dispatched_at,
                   s.station_name
            FROM dbo.WQ_Alert_Log al
            JOIN dbo.WQ_Indicators i ON i.indicator_id = al.indicator_id
            JOIN dbo.WQ_Readings   r ON r.reading_id   = al.reading_id
            JOIN dbo.WQ_Stations   s ON s.station_id   = r.station_id
            WHERE al.org_id = ?
            ORDER BY al.dispatched_at DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
            """,
            [org, offset, per_page]
        )
        logs = []
        if rows:
            for r in rows:
                logs.append({
                    'id': r[0], 'indicator': r[1], 'value': r[2],
                    'channel': r[3], 'recipient': r[4], 'status': r[5],
                    'dispatched_at': str(r[6]), 'station': r[7],
                })
        return jsonify({'success': True, 'logs': logs, 'total': total, 'page': page})
    except Exception as exc:
        logger.exception('WQ alert_log error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Piper Diagram ──────────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/piper', methods=['GET'])
@require_wq_auth()
def analytics_piper():
    """
    Return Piper diagram data for one or more stations / readings.
    Query: station_id (repeat for multiple), date_from, date_to
    Each reading event becomes one point.
    """
    org        = g.wq_org_id
    station_ids = request.args.getlist('station_id', type=int)
    date_from   = request.args.get('date_from')
    date_to     = request.args.get('date_to')

    where_clauses = ['r.org_id = ?']
    params        = [org]
    if station_ids:
        placeholders = ','.join(['?'] * len(station_ids))
        where_clauses.append(f'r.station_id IN ({placeholders})')
        params.extend(station_ids)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)
    where_sql = ' AND '.join(where_clauses)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, s.station_name,
                   i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Stations   s ON s.station_id   = r.station_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE {where_sql}
              AND i.indicator_code IN ('CA','MG','NA','K','HCO3','CO3','CL','SO4')
              AND m.measured_val IS NOT NULL
              AND m.is_censored = 0
              AND m.quality_flag = 'Approved'
            ORDER BY r.reading_date
            """,
            params
        )

        # Group by reading_id
        from collections import defaultdict
        readings_map = defaultdict(lambda: {'date': None, 'station': None, 'conc': {}})
        if rows:
            for row in rows:
                rid, rdate, sname, code, val = row
                readings_map[rid]['date']    = str(rdate)
                readings_map[rid]['station'] = sname
                readings_map[rid]['conc'][code] = val

        points = []
        for rid, info in readings_map.items():
            piper = compute_piper_data(info['conc'])
            points.append({
                'reading_id': rid,
                'date':       info['date'],
                'station':    info['station'],
                **piper,
            })

        return jsonify({'success': True, 'points': points})
    except Exception as exc:
        logger.exception('WQ analytics_piper error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Stiff Diagram ──────────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/stiff', methods=['GET'])
@require_wq_auth()
def analytics_stiff():
    """
    Return Stiff diagram polygon data for recent readings.
    Query: station_id (repeat), date_from, date_to, limit (max 10)
    """
    org         = g.wq_org_id
    station_ids = request.args.getlist('station_id', type=int)
    date_from   = request.args.get('date_from')
    date_to     = request.args.get('date_to')
    limit       = min(request.args.get('limit', 5, type=int), 10)

    where_clauses = ['r.org_id = ?']
    params        = [org]
    if station_ids:
        placeholders = ','.join(['?'] * len(station_ids))
        where_clauses.append(f'r.station_id IN ({placeholders})')
        params.extend(station_ids)
    if date_from:
        where_clauses.append('r.reading_date >= ?')
        params.append(date_from)
    if date_to:
        where_clauses.append('r.reading_date <= ?')
        params.append(date_to)
    where_sql = ' AND '.join(where_clauses)

    try:
        reading_rows = _q(
            f"""
            SELECT TOP {limit} r.reading_id, r.reading_date, s.station_name
            FROM dbo.WQ_Readings r
            JOIN dbo.WQ_Stations s ON s.station_id = r.station_id
            WHERE {where_sql}
            ORDER BY r.reading_date DESC
            """,
            params
        )
        if not reading_rows:
            return jsonify({'success': True, 'diagrams': []})

        reading_ids = [r[0] for r in reading_rows]
        reading_meta = {r[0]: {'date': str(r[1]), 'station': r[2]} for r in reading_rows}

        meas_rows = _q(
            f"""
            SELECT m.reading_id, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE m.reading_id IN ({','.join(['?']*len(reading_ids))})
              AND i.indicator_code IN ('CA','MG','NA','K','HCO3','CO3','CL','SO4')
              AND m.measured_val IS NOT NULL AND m.is_censored = 0
            """,
            reading_ids
        )

        from collections import defaultdict
        conc_map = defaultdict(dict)
        if meas_rows:
            for rid, code, val in meas_rows:
                conc_map[rid][code] = val

        diagrams = []
        for rid in reading_ids:
            meta = reading_meta[rid]
            label = f"{meta['station']} — {meta['date']}"
            stiff = compute_stiff_data(conc_map[rid], label=label)
            stiff['reading_id'] = rid
            stiff['date']       = meta['date']
            stiff['station']    = meta['station']
            diagrams.append(stiff)

        return jsonify({'success': True, 'diagrams': diagrams})
    except Exception as exc:
        logger.exception('WQ analytics_stiff error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Mann-Kendall Trend ─────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/mann-kendall', methods=['GET'])
@require_wq_auth()
def analytics_mann_kendall():
    """
    Mann-Kendall trend test for one indicator at one station.
    Query: station_id (required), indicator (code, default EC), date_from, date_to
    """
    org            = g.wq_org_id
    station_id     = request.args.get('station_id', type=int)
    indicator_code = request.args.get('indicator', 'EC')
    date_from      = request.args.get('date_from')
    date_to        = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org, indicator_code]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_date, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ? AND i.indicator_code = ?
              AND m.measured_val IS NOT NULL AND m.is_censored = 0
              AND m.quality_flag = 'Approved'
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        dates  = [r[0] for r in rows] if rows else []
        values = [r[1] for r in rows] if rows else []

        result = compute_mann_kendall(dates, values)
        result['indicator_code'] = indicator_code
        return jsonify({'success': True, 'mann_kendall': result})
    except Exception as exc:
        logger.exception('WQ analytics_mann_kendall error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Langelier Saturation Index ─────────────────────────────────────

@wq_bp.route('/api/wq/analytics/lsi', methods=['GET'])
@require_wq_auth()
def analytics_lsi():
    """
    LSI and RSI for each reading at a station.
    Query: station_id (required), date_from, date_to
    Requires: PH, TEMP, TDS (or EC→TDS conversion), CA, ALK measurements.
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND i.indicator_code IN ('PH','TEMP','TDS','EC','CA','ALK','HARD')
              AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        from collections import defaultdict
        reading_vals = defaultdict(dict)
        reading_dates = {}
        if rows:
            for rid, rdate, code, val in rows:
                reading_vals[rid][code] = val
                reading_dates[rid] = str(rdate)

        results = []
        for rid, vals in reading_vals.items():
            pH   = vals.get('PH')
            temp = vals.get('TEMP', 20.0)
            tds  = vals.get('TDS') or (vals.get('EC', 0) * 0.64 if vals.get('EC') else None)
            ca   = vals.get('CA')
            alk  = vals.get('ALK') or (vals.get('HARD', 0) * 0.5 if vals.get('HARD') else None)

            if pH is None or tds is None or ca is None or alk is None:
                results.append({
                    'reading_id': rid, 'date': reading_dates[rid],
                    'lsi': None, 'error': 'Missing required parameters (PH, TDS/EC, CA, ALK/HARD)'
                })
                continue

            lsi_data = compute_lsi(pH, temp or 20.0, tds, ca, alk)
            results.append({
                'reading_id': rid,
                'date':       reading_dates[rid],
                **lsi_data,
            })

        return jsonify({'success': True, 'lsi_series': results})
    except Exception as exc:
        logger.exception('WQ analytics_lsi error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Eutrophication ─────────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/eutrophication', methods=['GET'])
@require_wq_auth()
def analytics_eutrophication():
    """
    Carlson TSI for each reading at a station.
    Query: station_id (required), date_from, date_to
    Uses: TP (µg/L), CHLA (µg/L), SECCHI (m)
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND i.indicator_code IN ('TP','CHLA','SECCHI')
              AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        from collections import defaultdict
        rv = defaultdict(dict)
        rd = {}
        if rows:
            for rid, rdate, code, val in rows:
                rv[rid][code] = val
                rd[rid] = str(rdate)

        results = []
        for rid, vals in rv.items():
            tsi = compute_trophic_state_index(
                tp_ug_L    = vals.get('TP'),
                chl_a_ug_L = vals.get('CHLA'),
                secchi_m   = vals.get('SECCHI'),
            )
            results.append({'reading_id': rid, 'date': rd[rid], **tsi})

        return jsonify({'success': True, 'eutrophication': results})
    except Exception as exc:
        logger.exception('WQ analytics_eutrophication error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Inorganic Nitrogen (DWAF Vol 7 trophic bands) ──────────────────

# Conversion factors from "as ion" (how NH4/NO2/NO3 are stored in this app) to
# "as N" (how the DWAF Vol 7 guideline expresses inorganic nitrogen), based on
# molar mass ratios: N=14.007, NH4=18.04, NO2=46.01, NO3=62.00.
_ION_TO_N_FACTOR = {'NH4': 14.007 / 18.04, 'NO2': 14.007 / 46.01, 'NO3': 14.007 / 62.00}


def _inorganic_n_trophic_status(n_mgL):
    """DWAF Vol 7 (1996) trophic bands for inorganic nitrogen (mg N/L)."""
    if n_mgL is None:
        return None
    if n_mgL < 0.5:
        return 'oligotrophic'
    if n_mgL < 2.5:
        return 'mesotrophic'
    if n_mgL < 10:
        return 'eutrophic'
    return 'hypertrophic'


@wq_bp.route('/api/wq/analytics/inorganic-nitrogen', methods=['GET'])
@require_wq_auth()
def analytics_inorganic_nitrogen():
    """
    Derived Inorganic Nitrogen (NH4 + NO2 + NO3, converted to N-equivalent and
    summed) classified against DWAF Vol 7 (1996) trophic bands. Not a stored
    per-measurement compliance flag — this is a read-only, multi-indicator
    calculation, computed the same way compute_trophic_state_index() combines
    TP/CHLA/SECCHI for analytics_eutrophication() above.
    Query: station_id (required), date_from, date_to
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND i.indicator_code IN ('NH4','NO2','NO3')
              AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        from collections import defaultdict
        rv = defaultdict(dict)
        rd = {}
        if rows:
            for rid, rdate, code, val in rows:
                rv[rid][code] = val
                rd[rid] = str(rdate)

        results = []
        for rid, vals in rv.items():
            present = [code for code in ('NH4', 'NO2', 'NO3') if code in vals]
            if not present:
                continue
            n_total = sum(vals[code] * _ION_TO_N_FACTOR[code] for code in present)
            results.append({
                'reading_id': rid,
                'date': rd[rid],
                'inorganic_n_mgL': round(n_total, 4),
                'trophic_status': _inorganic_n_trophic_status(n_total),
                'components_present': present,
                'reference': 'DWAF Vol 7 (1996): <0.5 oligotrophic, 0.5-2.5 mesotrophic, 2.5-10 eutrophic, >10 hypertrophic (mg N/L)',
            })

        return jsonify({'success': True, 'inorganic_nitrogen': results})
    except Exception as exc:
        logger.exception('WQ analytics_inorganic_nitrogen error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: AMD Assessment ─────────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/amd', methods=['GET'])
@require_wq_auth()
def analytics_amd():
    """
    Acid Mine Drainage (AMD) indicators for a station.
    Flags readings as AMD-suspect based on SA Witwatersrand criteria.
    pH ≤ 4, EC > 500 mS/m, SO4 > 1000 mg/L, Fe(total) > 10 mg/L
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND i.indicator_code IN ('PH','EC','SO4','FE','FE2','FE3','TOTAL_ACID','MN','AS')
              AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        from collections import defaultdict
        rv = defaultdict(dict)
        rd = {}
        if rows:
            for rid, rdate, code, val in rows:
                rv[rid][code] = val
                rd[rid] = str(rdate)

        results = []
        for rid, vals in rv.items():
            amd = compute_amd_flags(vals)
            results.append({
                'reading_id':  rid,
                'date':        rd[rid],
                'parameters':  vals,
                'amd_flags':   amd['flags'],
                'amd_risk':    amd['amd_risk'],
                'is_amd_suspect': amd['is_amd_suspect'],
            })

        return jsonify({'success': True, 'amd': results})
    except Exception as exc:
        logger.exception('WQ analytics_amd error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Seasonal Comparison ───────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/seasonal', methods=['GET'])
@require_wq_auth()
def analytics_seasonal():
    """
    Compare mean values per indicator between Wet and Dry seasons.
    SA Wet season = Oct–Apr, Dry = May–Sep.
    Query: station_id (required), date_from, date_to
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT i.indicator_code, i.indicator_name, i.unit,
                   CASE WHEN MONTH(r.reading_date) IN (10,11,12,1,2,3,4)
                        THEN 'Wet' ELSE 'Dry' END AS season,
                   AVG(m.measured_val)  AS mean_val,
                   MIN(m.measured_val)  AS min_val,
                   MAX(m.measured_val)  AS max_val,
                   COUNT(m.meas_id)     AS n
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND m.measured_val IS NOT NULL
              AND m.quality_flag IN ('Approved', 'Estimated')
              {where_extra}
            GROUP BY i.indicator_code, i.indicator_name, i.unit,
                     CASE WHEN MONTH(r.reading_date) IN (10,11,12,1,2,3,4)
                          THEN 'Wet' ELSE 'Dry' END
            ORDER BY i.display_order
            """,
            params
        )

        from collections import defaultdict
        by_indicator = defaultdict(lambda: {'wet': None, 'dry': None, 'name': None, 'unit': None})
        if rows:
            for code, name, unit, season, mean, mn, mx, n in rows:
                by_indicator[code]['name'] = name
                by_indicator[code]['unit'] = unit
                entry = {'mean': round(mean, 4) if mean is not None else None,
                         'min': round(mn, 4) if mn is not None else None,
                         'max': round(mx, 4) if mx is not None else None,
                         'n': n}
                if season == 'Wet':
                    by_indicator[code]['wet'] = entry
                else:
                    by_indicator[code]['dry'] = entry

        results = [
            {'code': code, 'name': v['name'], 'unit': v['unit'],
             'wet': v['wet'], 'dry': v['dry']}
            for code, v in by_indicator.items()
        ]

        return jsonify({'success': True, 'seasonal': results})
    except Exception as exc:
        logger.exception('WQ analytics_seasonal error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Pollution Load ─────────────────────────────────────────────────

@wq_bp.route('/api/wq/analytics/pollution-load', methods=['GET'])
@require_wq_auth()
def analytics_pollution_load():
    """
    Calculate pollution load (kg/day) for each reading that has a flow rate.
    Load = concentration (mg/L) × flow (m³/s) × 86.4
    Query: station_id (required), indicator (code, default TDS), date_from, date_to
    """
    org            = g.wq_org_id
    station_id     = request.args.get('station_id', type=int)
    indicator_code = request.args.get('indicator', 'TDS')
    date_from      = request.args.get('date_from')
    date_to        = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org, indicator_code]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, r.flow_rate_m3s,
                   m.measured_val, i.unit
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ?
              AND i.indicator_code = ?
              AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """,
            params
        )

        results = []
        if rows:
            for rid, rdate, flow, val, unit in rows:
                load_kg_day = None
                if flow is not None and flow > 0 and val is not None:
                    load_kg_day = round(val * flow * 86.4, 2)
                results.append({
                    'reading_id':   rid,
                    'date':         str(rdate),
                    'concentration': val,
                    'unit':          unit,
                    'flow_m3s':     flow,
                    'load_kg_day':  load_kg_day,
                })

        return jsonify({
            'success':        True,
            'indicator_code': indicator_code,
            'series':         results,
            'note':           'Load (kg/day) = Concentration (mg/L) × Flow (m³/s) × 86.4. Requires flow_rate_m3s on each reading.',
        })
    except Exception as exc:
        logger.exception('WQ analytics_pollution_load error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ============================================================================
# Water Safety Plan — Risk Register + Corrective Actions
# ============================================================================
# WHO/Blue-Drop-style risk management: a documented hazard -> control measure
# -> monitoring register, plus tracking of non-conformances to closure. Not
# tied to any single compliance standard — usable alongside either SANS 241
# or TWQR.

@wq_bp.route('/api/wq/risk-register', methods=['GET'])
@require_wq_auth()
def list_risk_register():
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    status     = request.args.get('status')

    where = ['org_id = ?']
    params = [org]
    if station_id:
        where.append('station_id = ?')
        params.append(station_id)
    if status:
        where.append('status = ?')
        params.append(status)

    try:
        rows = _q(
            f"""
            SELECT risk_id, station_id, linked_indicator_id, hazard_description, hazard_source,
                   likelihood, severity, risk_score, control_measure, monitoring_frequency,
                   responsible_person, status, review_date, created_at, updated_at
            FROM dbo.WQ_Risk_Register
            WHERE {' AND '.join(where)}
            ORDER BY risk_score DESC, created_at DESC
            """, params
        )
        risks = [
            {
                'risk_id': r[0], 'station_id': r[1], 'linked_indicator_id': r[2],
                'hazard_description': r[3], 'hazard_source': r[4],
                'likelihood': r[5], 'severity': r[6], 'risk_score': r[7],
                'control_measure': r[8], 'monitoring_frequency': r[9],
                'responsible_person': r[10], 'status': r[11],
                'review_date': str(r[12]) if r[12] else None,
                'created_at': str(r[13]), 'updated_at': str(r[14]),
            }
            for r in (rows or [])
        ]
        return jsonify({'success': True, 'risks': risks})
    except Exception as exc:
        logger.exception('WQ list_risk_register error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/risk-register', methods=['POST'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def create_risk_register():
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}
    hazard_description = (data.get('hazard_description') or '').strip()
    try:
        likelihood = int(data.get('likelihood'))
        severity   = int(data.get('severity'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'likelihood and severity are required integers (1-5).'}), 400
    if not hazard_description or not (1 <= likelihood <= 5) or not (1 <= severity <= 5):
        return jsonify({'success': False, 'error': 'hazard_description is required; likelihood/severity must be 1-5.'}), 400

    try:
        _q(
            """
            INSERT INTO dbo.WQ_Risk_Register
                (org_id, station_id, linked_indicator_id, hazard_description, hazard_source,
                 likelihood, severity, control_measure, monitoring_frequency, responsible_person, review_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            [org, data.get('station_id'), data.get('linked_indicator_id'), hazard_description,
             data.get('hazard_source'), likelihood, severity, data.get('control_measure'),
             data.get('monitoring_frequency'), data.get('responsible_person'), data.get('review_date')],
            fetch=False, commit=True
        )
        new_row = _q(
            "SELECT TOP 1 risk_id FROM dbo.WQ_Risk_Register WHERE org_id=? ORDER BY risk_id DESC", [org]
        )
        return jsonify({'success': True, 'risk_id': new_row[0][0] if new_row else None}), 201
    except Exception as exc:
        logger.exception('WQ create_risk_register error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/risk-register/<int:risk_id>', methods=['PUT'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def update_risk_register(risk_id):
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}
    fields, params = [], []
    for key, col in (('hazard_description', 'hazard_description'), ('hazard_source', 'hazard_source'),
                      ('control_measure', 'control_measure'), ('monitoring_frequency', 'monitoring_frequency'),
                      ('responsible_person', 'responsible_person'), ('review_date', 'review_date')):
        if key in data:
            fields.append(f'{col} = ?')
            params.append(data[key])
    if 'likelihood' in data:
        fields.append('likelihood = ?'); params.append(int(data['likelihood']))
    if 'severity' in data:
        fields.append('severity = ?'); params.append(int(data['severity']))
    if 'status' in data and data['status'] in ('open', 'mitigated', 'closed'):
        fields.append('status = ?'); params.append(data['status'])
    if not fields:
        return jsonify({'success': False, 'error': 'No updatable fields provided.'}), 400
    fields.append('updated_at = GETDATE()')
    params.extend([risk_id, org])

    try:
        _q(f"UPDATE dbo.WQ_Risk_Register SET {', '.join(fields)} WHERE risk_id=? AND org_id=?",
           params, fetch=False, commit=True)
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ update_risk_register error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/risk-register/<int:risk_id>', methods=['DELETE'])
@require_wq_auth(roles=['admin', 'standard_user'])
def delete_risk_register(risk_id):
    org = g.wq_org_id
    try:
        _q("DELETE FROM dbo.WQ_Risk_Register WHERE risk_id=? AND org_id=?", [risk_id, org], fetch=False, commit=True)
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ delete_risk_register error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/corrective-actions', methods=['GET'])
@require_wq_auth()
def list_corrective_actions():
    org    = g.wq_org_id
    status = request.args.get('status')
    where  = ['org_id = ?']
    params = [org]
    if status:
        where.append('status = ?')
        params.append(status)

    try:
        rows = _q(
            f"""
            SELECT action_id, station_id, indicator_id, reading_id, description, root_cause,
                   action_taken, assigned_to, due_date, status, auto_generated, closed_at, created_at
            FROM dbo.WQ_Corrective_Actions
            WHERE {' AND '.join(where)}
            ORDER BY created_at DESC
            """, params
        )
        actions = [
            {
                'action_id': r[0], 'station_id': r[1], 'indicator_id': r[2], 'reading_id': r[3],
                'description': r[4], 'root_cause': r[5], 'action_taken': r[6], 'assigned_to': r[7],
                'due_date': str(r[8]) if r[8] else None, 'status': r[9], 'auto_generated': bool(r[10]),
                'closed_at': str(r[11]) if r[11] else None, 'created_at': str(r[12]),
            }
            for r in (rows or [])
        ]
        return jsonify({'success': True, 'actions': actions})
    except Exception as exc:
        logger.exception('WQ list_corrective_actions error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/corrective-actions', methods=['POST'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def create_corrective_action():
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}
    description = (data.get('description') or '').strip()
    if not description:
        return jsonify({'success': False, 'error': 'description is required.'}), 400
    try:
        _q(
            """
            INSERT INTO dbo.WQ_Corrective_Actions
                (org_id, station_id, indicator_id, reading_id, description, root_cause,
                 action_taken, assigned_to, due_date)
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            [org, data.get('station_id'), data.get('indicator_id'), data.get('reading_id'), description,
             data.get('root_cause'), data.get('action_taken'), data.get('assigned_to'), data.get('due_date')],
            fetch=False, commit=True
        )
        new_row = _q(
            "SELECT TOP 1 action_id FROM dbo.WQ_Corrective_Actions WHERE org_id=? ORDER BY action_id DESC", [org]
        )
        return jsonify({'success': True, 'action_id': new_row[0][0] if new_row else None}), 201
    except Exception as exc:
        logger.exception('WQ create_corrective_action error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/corrective-actions/<int:action_id>', methods=['PUT'])
@require_wq_auth(roles=['admin', 'analyst', 'standard_user'])
def update_corrective_action(action_id):
    org  = g.wq_org_id
    data = request.get_json(force=True) or {}
    fields, params = [], []
    for key in ('description', 'root_cause', 'action_taken', 'assigned_to', 'due_date'):
        if key in data:
            fields.append(f'{key} = ?')
            params.append(data[key])
    if 'status' in data and data['status'] in ('open', 'in_progress', 'closed', 'overdue'):
        fields.append('status = ?')
        params.append(data['status'])
        if data['status'] == 'closed':
            fields.append('closed_at = GETDATE()')
    if not fields:
        return jsonify({'success': False, 'error': 'No updatable fields provided.'}), 400
    params.extend([action_id, org])

    try:
        _q(f"UPDATE dbo.WQ_Corrective_Actions SET {', '.join(fields)} WHERE action_id=? AND org_id=?",
           params, fetch=False, commit=True)
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ update_corrective_action error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/corrective-actions/<int:action_id>', methods=['DELETE'])
@require_wq_auth(roles=['admin', 'standard_user'])
def delete_corrective_action(action_id):
    org = g.wq_org_id
    try:
        _q("DELETE FROM dbo.WQ_Corrective_Actions WHERE action_id=? AND org_id=?",
           [action_id, org], fetch=False, commit=True)
        return jsonify({'success': True})
    except Exception as exc:
        logger.exception('WQ delete_corrective_action error')
        return jsonify({'success': False, 'error': str(exc)}), 500


@wq_bp.route('/api/wq/wsp/export', methods=['GET'])
@require_wq_auth()
def wsp_export():
    """Excel export of the risk register + open corrective actions."""
    org = g.wq_org_id
    try:
        risk_rows = _q(
            """
            SELECT hazard_description, hazard_source, likelihood, severity, risk_score,
                   control_measure, monitoring_frequency, responsible_person, status, review_date
            FROM dbo.WQ_Risk_Register WHERE org_id=? ORDER BY risk_score DESC
            """, [org]
        ) or []
        action_rows = _q(
            """
            SELECT description, root_cause, action_taken, assigned_to, due_date, status, auto_generated, created_at
            FROM dbo.WQ_Corrective_Actions WHERE org_id=? AND status != 'closed' ORDER BY created_at DESC
            """, [org]
        ) or []

        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Risk Register'
        header_fill = PatternFill('solid', fgColor='0D9488')
        header_font = Font(color='FFFFFF', bold=True, size=10)

        risk_headers = ['Hazard', 'Source', 'Likelihood', 'Severity', 'Risk Score',
                         'Control Measure', 'Monitoring Frequency', 'Responsible', 'Status', 'Review Date']
        for ci, h in enumerate(risk_headers, start=1):
            cell = ws1.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for ri, row in enumerate(risk_rows, start=2):
            for ci, val in enumerate(row, start=1):
                ws1.cell(row=ri, column=ci, value=str(val) if val is not None else '')

        ws2 = wb.create_sheet(title='Corrective Actions (Open)')
        action_headers = ['Description', 'Root Cause', 'Action Taken', 'Assigned To',
                           'Due Date', 'Status', 'Auto-Generated', 'Created']
        for ci, h in enumerate(action_headers, start=1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for ri, row in enumerate(action_rows, start=2):
            for ci, val in enumerate(row, start=1):
                ws2.cell(row=ri, column=ci, value=str(val) if val is not None else '')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='water_safety_plan_export.xlsx',
        )
    except Exception as exc:
        logger.exception('WQ wsp_export error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Multi-Use Compliance (Irrigation / Livestock / Aquaculture) ────

@wq_bp.route('/api/wq/analytics/multi-use-compliance', methods=['GET'])
@require_wq_auth()
def analytics_multi_use_compliance():
    """
    Irrigation (DWAF Vol 4 SAR/EC), livestock watering (DWAF Vol 5 TDS-per-
    species) and aquaculture (DWAF Vol 6) fitness-for-use verdicts per
    reading — derived live from already-stored measurements, never persisted
    (same pattern as analytics_amd()).
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    if not station_id:
        return jsonify({'success': False, 'error': 'station_id is required.'}), 400

    where_extra = ''
    params = [station_id, org]
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, r.reading_date, i.indicator_code, m.measured_val
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.station_id = ? AND r.org_id = ? AND m.measured_val IS NOT NULL
              {where_extra}
            ORDER BY r.reading_date
            """, params
        )

        from collections import defaultdict
        rv = defaultdict(dict)
        rd = {}
        if rows:
            for rid, rdate, code, val in rows:
                rv[rid][code] = val
                rd[rid] = str(rdate)

        livestock_bands = _fetch_livestock_bands()
        aqua_criteria   = _fetch_aquaculture_criteria()

        irrigation, livestock, aquaculture = [], [], []
        for rid, vals in rv.items():
            irr = compute_irrigation_suitability(vals, vals.get('EC'))
            irrigation.append({'reading_id': rid, 'date': rd[rid], **irr})

            live = compute_livestock_suitability(vals.get('TDS'), livestock_bands)
            livestock.append({'reading_id': rid, 'date': rd[rid], **live})

            aqua = compute_aquaculture_suitability(vals, vals.get('HARD'), aqua_criteria)
            aquaculture.append({'reading_id': rid, 'date': rd[rid], **aqua})

        return jsonify({'success': True, 'irrigation': irrigation, 'livestock': livestock, 'aquaculture': aquaculture})
    except Exception as exc:
        logger.exception('WQ multi_use_compliance error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Analytics: Aquatic Ecosystem Health Score ──────────────────────────────────

@wq_bp.route('/api/wq/analytics/ecosystem-health', methods=['GET'])
@require_wq_auth()
def analytics_ecosystem_health():
    """
    Composite 0-100 Aquatic Ecosystem Health Score built on already-stored
    TWQR exceedance tiers (is_compliant_twqr/flag_twqr + twqr_cev/twqr_aev),
    aggregated over the given station/date range — same query-aggregation
    pattern as wq_blue_drop().
    """
    org        = g.wq_org_id
    station_id = request.args.get('station_id', type=int)
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')

    where_extra = ''
    params = [org]
    if station_id:
        where_extra += ' AND r.station_id = ?'
        params.append(station_id)
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)

    try:
        rows = _q(
            f"""
            SELECT r.reading_id, i.indicator_id, i.indicator_code, i.indicator_name,
                   m.measured_val, i.twqr_lower, i.twqr_upper, i.twqr_cev, i.twqr_aev,
                   i.twqr_basis, m.is_compliant_twqr
            FROM dbo.WQ_Measurements m
            JOIN dbo.WQ_Readings   r ON r.reading_id   = m.reading_id
            JOIN dbo.WQ_Indicators i ON i.indicator_id = m.indicator_id
            WHERE r.org_id = ? AND m.measured_val IS NOT NULL AND m.quality_flag = 'Approved'
              {where_extra}
            """, params
        )

        if not rows:
            return jsonify({'success': True, 'ecosystem_health': compute_ecosystem_health_score([])})

        hardness_by_reading = {}
        for rid, ind_id, code, name, val, lower, upper, cev, aev, basis, compliant in rows:
            if code == 'HARD':
                hardness_by_reading[rid] = val

        hardness_bands = _fetch_hardness_bands(org)

        measurements = []
        for rid, ind_id, code, name, val, lower, upper, cev, aev, basis, compliant in rows:
            if basis == 'hardness_dependent':
                band = classify_hardness(hardness_by_reading.get(rid))
                row_band = hardness_bands.get(ind_id, {}).get(band)
                if row_band:
                    upper, cev, aev = row_band
            measurements.append({
                'indicator_code': code, 'indicator_name': name, 'measured_val': val,
                'twqr_lower': lower, 'twqr_upper': upper, 'twqr_cev': cev, 'twqr_aev': aev,
                'twqr_basis': basis, 'is_compliant_twqr': bool(compliant) if compliant is not None else None,
            })

        result = compute_ecosystem_health_score(measurements)
        return jsonify({'success': True, 'ecosystem_health': result})
    except Exception as exc:
        logger.exception('WQ ecosystem_health error')
        return jsonify({'success': False, 'error': str(exc)}), 500


# ── Diagnostics: Contamination Source-Attribution Findings ─────────────────────

@wq_bp.route('/api/wq/diagnostics', methods=['GET'])
@require_wq_auth()
def wq_diagnostics():
    """Read-only view of persisted WQ_Diagnostic_Findings (computed at
    upload time by diagnose_contamination_source())."""
    org            = g.wq_org_id
    station_id     = request.args.get('station_id', type=int)
    reading_id     = request.args.get('reading_id', type=int)
    date_from      = request.args.get('date_from')
    date_to        = request.args.get('date_to')
    min_confidence = request.args.get('min_confidence', type=float)

    where_extra = ''
    params = [org]
    if station_id:
        where_extra += ' AND f.station_id = ?'
        params.append(station_id)
    if reading_id:
        where_extra += ' AND f.reading_id = ?'
        params.append(reading_id)
    if date_from:
        where_extra += ' AND r.reading_date >= ?'
        params.append(date_from)
    if date_to:
        where_extra += ' AND r.reading_date <= ?'
        params.append(date_to)
    if min_confidence is not None:
        where_extra += ' AND f.confidence_pct >= ?'
        params.append(min_confidence)

    try:
        rows = _q(
            f"""
            SELECT f.finding_id, f.reading_id, r.reading_date, s.station_name,
                   f.source_code, f.source_label, f.confidence_pct,
                   f.matched_signals, f.contradicting_signals, f.contextual_modifiers_applied,
                   f.explanation_text, f.created_at
            FROM dbo.WQ_Diagnostic_Findings f
            JOIN dbo.WQ_Readings r ON r.reading_id = f.reading_id
            JOIN dbo.WQ_Stations s ON s.station_id = f.station_id
            WHERE f.org_id = ?
              {where_extra}
            ORDER BY r.reading_date DESC, f.confidence_pct DESC
            """, params
        )

        findings = []
        for row in (rows or []):
            findings.append({
                'finding_id':            row[0],
                'reading_id':            row[1],
                'date':                  str(row[2]),
                'station':               row[3],
                'source_code':           row[4],
                'source_label':          row[5],
                'confidence_pct':        row[6],
                'matched_signals':       json.loads(row[7]) if row[7] else [],
                'contradicting_signals': json.loads(row[8]) if row[8] else [],
                'contextual_modifiers':  json.loads(row[9]) if row[9] else [],
                'explanation_text':      row[10],
                'created_at':            str(row[11]),
            })
        return jsonify({'success': True, 'findings': findings})
    except Exception as exc:
        logger.exception('WQ diagnostics error')
        return jsonify({'success': False, 'error': str(exc)}), 500
