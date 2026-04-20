"""
Converts the new "Main" format Excel files (Assegai, Lower Komati, Ngwempisi, Upper Komati)
into separate per-parameter files that the existing upload system accepts.

New format: 3-row header, all 3 parameters side-by-side in one sheet
Old format: 1-row header, one parameter per file

Run from the project root:
    python convert_main_files.py
"""

import os
import openpyxl
import pandas as pd
from datetime import datetime

# Input files and their catchment names (must match what's in the database exactly)
SOURCE_FILES = {
    "Examples/Assegai(Main).xlsx":        "Assegai",
    "Examples/Lower Komati (Main).xlsx":  "Lower Komati",
    "Examples/Ngwempisi(Main).xlsx":      "Ngwempisi",
    "Examples/Upper Komati (Main).xlsx":  "Upper Komati",
}

# New-format files: 3-row header, avg/stdev embedded per data row, fixed column positions
# Recharge: cols 0(date),1(value),3(avg),4(stdev),5(xi-xm),6(SDI)
# Baseflow:  cols 8(date),9(value),10(avg),11(stdev),12(xi-xm),13(SDI)
# GWL:       cols 15(date),16(value),17(avg),18(stdev),19(xi-xm),20(SDI)
NEW_FORMAT_FILES = {
    "Examples/Crocodile (Main).xlsx": "Crocodile",
    "Examples/Sabie.xlsx":            "Sabie",
    "Examples/Sand (Main).xlsx":      "Sand",
}

OUTPUT_DIR = "Examples/converted"


def detect_parameter(section_title):
    """Identify which parameter a block represents from its section title."""
    if section_title is None:
        return None
    t = str(section_title).lower()
    if "recharge" in t:
        return "recharge"
    if "baseflow" in t:
        return "baseflow"
    if "groundwater" in t or "gwl" in t or "gw level" in t or "ground water" in t:
        return "gwlevel"
    return None


def parse_main_file(filepath):
    """
    Parse a *Main*.xlsx file and return a dict:
      { 'recharge': DataFrame, 'baseflow': DataFrame, 'gwlevel': DataFrame }
    each with columns matching what the upload system expects.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    all_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 4:
        raise ValueError(f"File has fewer than 4 rows: {filepath}")

    row0 = all_rows[0]   # Section titles: 'Drought index - Baseflow', 'RECHARGE', etc.
    row1 = all_rows[1]   # Average/StDev values
    row2 = all_rows[2]   # Column headers: 'Date', 'value col', 'Xi - Xm', 'SDI', None, ...
    data_rows = all_rows[3:]  # Actual data

    # Find where each parameter block starts by locating 'Date' columns in row2
    # Blocks are separated by a None column; each block is 4 columns wide
    block_starts = [i for i, v in enumerate(row2) if v is not None and str(v).strip().lower() == "date"]

    if len(block_starts) < 3:
        # Fallback: assume fixed positions 0, 5, 10
        block_starts = [0, 5, 10]

    result = {}

    for start_col in block_starts:
        # Identify parameter from row0 section title
        section_title = row0[start_col] if start_col < len(row0) else None
        # Also check neighbouring cells since the title may be in the first cell
        for offset in range(5):
            if start_col + offset < len(row0) and row0[start_col + offset] is not None:
                detected = detect_parameter(row0[start_col + offset])
                if detected:
                    param = detected
                    break
        else:
            continue  # Could not detect parameter type for this block

        # Stats from row1
        avg_val  = row1[start_col + 1] if (start_col + 1) < len(row1) else None
        stdev_val = row1[start_col + 3] if (start_col + 3) < len(row1) else None

        # Extract data
        dates, values, xi_xm, sdi = [], [], [], []
        for row in data_rows:
            d    = row[start_col]     if start_col < len(row) else None
            v    = row[start_col + 1] if (start_col + 1) < len(row) else None
            dev  = row[start_col + 2] if (start_col + 2) < len(row) else None
            s    = row[start_col + 3] if (start_col + 3) < len(row) else None

            if d is None:
                continue
            try:
                parsed_date = pd.to_datetime(d)
            except Exception:
                continue

            dates.append(parsed_date)
            values.append(v)
            xi_xm.append(dev)
            sdi.append(s)

        if not dates:
            continue

        # Build DataFrame with column names the upload system recognises
        if param == "baseflow":
            df = pd.DataFrame({
                "measurement date":    dates,
                "baseflow value":      values,
                "average baseflow":    [avg_val] * len(dates),
                "stdev":               [stdev_val] * len(dates),
                "standardized baseflow": sdi,
                "baseflow deviation":  xi_xm,
            })

        elif param == "recharge":
            df = pd.DataFrame({
                "measurement date":        dates,
                "recharge (inches)":       values,   # value is mm but column name must match
                "average recharge":        [avg_val] * len(dates),
                "stdev":                   [stdev_val] * len(dates),
                "drought index - recharge": sdi,
                "recharge deviation":      xi_xm,
            })

        elif param == "gwlevel":
            df = pd.DataFrame({
                "measurement date":    dates,
                "gw level":            values,
                "average gw level":    [avg_val] * len(dates),
                "stdev":               [stdev_val] * len(dates),
                "standardized gw level": sdi,
                "gw level deviation":  xi_xm,
            })

        result[param] = df

    return result


def parse_new_format_file(filepath):
    """
    Parse Sabie/Crocodile/Sand style files:
      - Row 0: location label (ignored)
      - Row 1: parameter names (Recharge / Baseflow / Groundwater levels)
      - Row 2: column headers
      - Row 3+: data rows with avg & stdev embedded as columns

    Fixed layout (0-indexed):
      Recharge  → date=0, val=1, avg=3, stdev=4, xi_xm=5, sdi=6
      Baseflow  → date=8, val=9, avg=10, stdev=11, xi_xm=12, sdi=13
      GWL       → date=15, val=16, avg=17, stdev=18, xi_xm=19, sdi=20
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    all_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 4:
        raise ValueError(f"File has fewer than 4 rows: {filepath}")

    data_rows = all_rows[3:]   # skip 3-row header

    BLOCKS = {
        "recharge": {"date": 0, "val": 1, "avg": 3, "stdev": 4, "xi_xm": 5, "sdi": 6},
        "baseflow": {"date": 8, "val": 9, "avg": 10, "stdev": 11, "xi_xm": 12, "sdi": 13},
        "gwlevel":  {"date": 15, "val": 16, "avg": 17, "stdev": 18, "xi_xm": 19, "sdi": 20},
    }

    def _get(row, idx):
        return row[idx] if idx < len(row) else None

    result = {}

    for param, cols in BLOCKS.items():
        dates, values, avgs, stdevs, xi_xm_list, sdi_list = [], [], [], [], [], []

        for row in data_rows:
            d = _get(row, cols["date"])
            if d is None:
                continue
            try:
                parsed_date = pd.to_datetime(d)
            except Exception:
                continue

            dates.append(parsed_date)
            values.append(_get(row, cols["val"]))
            avgs.append(_get(row, cols["avg"]))
            stdevs.append(_get(row, cols["stdev"]))
            xi_xm_list.append(_get(row, cols["xi_xm"]))
            sdi_list.append(_get(row, cols["sdi"]))

        if not dates:
            continue

        if param == "baseflow":
            df = pd.DataFrame({
                "measurement date":      dates,
                "baseflow value":        values,
                "average baseflow":      avgs,
                "stdev":                 stdevs,
                "standardized baseflow": sdi_list,
                "baseflow deviation":    xi_xm_list,
            })
        elif param == "recharge":
            df = pd.DataFrame({
                "measurement date":         dates,
                "recharge (inches)":        values,
                "average recharge":         avgs,
                "stdev":                    stdevs,
                "drought index - recharge": sdi_list,
                "recharge deviation":       xi_xm_list,
            })
        elif param == "gwlevel":
            df = pd.DataFrame({
                "measurement date":      dates,
                "gw level":              values,
                "average gw level":      avgs,
                "stdev":                 stdevs,
                "standardized gw level": sdi_list,
                "gw level deviation":    xi_xm_list,
            })

        result[param] = df

    return result


def convert_all():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    manifest = []   # List of (output_file, catchment, category) for the user

    for filepath, catchment in SOURCE_FILES.items():
        print(f"\nProcessing: {filepath}  →  catchment: {catchment}")
        try:
            param_data = parse_main_file(filepath)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        for category, df in param_data.items():
            safe_catchment = catchment.replace(" ", "_")
            out_name = f"{safe_catchment}_{category}.xlsx"
            out_path = os.path.join(OUTPUT_DIR, out_name)

            df.to_excel(out_path, index=False)
            print(f"  ✓ {category}: {len(df)} rows  →  {out_path}")
            manifest.append((out_path, catchment, category))

    for filepath, catchment in NEW_FORMAT_FILES.items():
        print(f"\nProcessing (new format): {filepath}  →  catchment: {catchment}")
        try:
            param_data = parse_new_format_file(filepath)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        for category, df in param_data.items():
            safe_catchment = catchment.replace(" ", "_")
            out_name = f"{safe_catchment}_{category}.xlsx"
            out_path = os.path.join(OUTPUT_DIR, out_name)

            df.to_excel(out_path, index=False)
            print(f"  ✓ {category}: {len(df)} rows  →  {out_path}")
            manifest.append((out_path, catchment, category))

    print("\n" + "=" * 65)
    print("CONVERSION COMPLETE")
    print("=" * 65)
    print(f"Output folder: {os.path.abspath(OUTPUT_DIR)}")
    print()
    print("Upload each file below via the dashboard (Upload Data tab):")
    print(f"{'File':<45} {'Catchment':<15} {'Category'}")
    print("-" * 65)
    for out_path, catchment, category in sorted(manifest):
        print(f"{os.path.basename(out_path):<45} {catchment:<15} {category}")

    print()
    print("Upload order doesn't matter. Use exact catchment names shown above.")


if __name__ == "__main__":
    convert_all()
